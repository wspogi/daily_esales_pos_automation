"""
============================================================
Daily E-Sales Report POS Automation
============================================================

Purpose:
    Connect to the POS SQL Server database, summarize the Daily
    E-Sales report for a selected period, and export it to Excel
    using the same layout as the uploaded APR.xls template.

Report layout based on APR.xls:
    Row 1  : Company name
    Row 2  : Operated by
    Row 3  : Report title
    Row 4  : Period covered
    Row 7  : Table headers
    Row 8+ : Terminal/POS rows
    Last   : Grand Total row

Main idea:
    Static POS/BIR/header details are from .env:
        - POS_BRANCH
        - POS_COMPANY_NAME
        - POS_OPERATOR_NAME
        - POS_TIN
        - POS_SERIAL_NO
        - POS_MIN
        - POS_PERMIT_NO

    Sales values are from SQL:
        - LastInvoiceNo
        - NetSalesWithVAT
        - ZeroRated
        - VATExempt
        - Vatable
        - TotalSalesNetOfVAT
        - OutputVAT

Date options:
    DATE_MODE=PREVIOUS_MONTH
    DATE_MODE=CURRENT_MONTH
    DATE_MODE=CUSTOM

Manual run override:
    py main.py 2026-04

Example:
    py main.py 2026-04
    This generates report for 04/01/2026 to 04/30/2026.

Important SQL note:
    sql/query_daily_esales.sql can work in two modes:

    1. No parameter markers:
        SELECT ...;

       Python will run it without passing dates.

    2. With two parameter markers:
        WHERE SaleDate >= ?
          AND SaleDate < ?

       Python will pass:
            start_date
            end_date_exclusive

============================================================
"""

import logging
import os
import sys
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import pyodbc
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ============================================================
# BASE PATHS
# ============================================================

BASE_DIR = Path(__file__).resolve().parent
ENV_PATH = BASE_DIR / ".env"


# ============================================================
# HELPER: PATH HANDLING
# ============================================================

def resolve_project_path(path_value: str, default_folder: str) -> Path:
    """
    Convert folder path from .env into a real Path.

    Why:
        If OUTPUT_DIR=output, dapat nasa project folder siya.
        If full path ang nilagay, gagamitin niya yung full path.
    """
    clean_value = (path_value or default_folder).strip()
    folder_path = Path(clean_value)

    if folder_path.is_absolute():
        return folder_path

    return BASE_DIR / folder_path


# ============================================================
# SETTINGS
# ============================================================

def load_settings() -> dict:
    """
    Load .env values.

    Why:
        Para hindi hardcoded ang SQL credentials, branch details,
        company name, POS permit details, output folder, and logs folder.

        Safe din for GitHub because actual .env should be ignored.
        Only .env.example should be committed.
    """
    load_dotenv(ENV_PATH)

    settings = {
        # SQL connection
        "SQL_SERVER": os.getenv("SQL_SERVER", "").strip(),
        "SQL_DATABASE": os.getenv("SQL_DATABASE", "").strip(),
        "SQL_USERNAME": os.getenv("SQL_USERNAME", "").strip(),
        "SQL_PASSWORD": os.getenv("SQL_PASSWORD", ""),
        "SQL_DRIVER": os.getenv("SQL_DRIVER", "ODBC Driver 17 for SQL Server").strip(),
        "SQL_ENCRYPT": os.getenv("SQL_ENCRYPT", "no").strip(),
        "SQL_TRUST_SERVER_CERTIFICATE": os.getenv("SQL_TRUST_SERVER_CERTIFICATE", "yes").strip(),

        # POS branch / report header info
        "POS_BRANCH": os.getenv("POS_BRANCH", "").strip(),
        "POS_COMPANY_NAME": os.getenv("POS_COMPANY_NAME", "").strip(),
        "POS_OPERATOR_NAME": os.getenv("POS_OPERATOR_NAME", "").strip(),
        "POS_TIN": os.getenv("POS_TIN", "").strip(),
        "POS_SERIAL_NO": os.getenv("POS_SERIAL_NO", "").strip(),
        "POS_MIN": os.getenv("POS_MIN", "").strip(),
        "POS_PERMIT_NO": os.getenv("POS_PERMIT_NO", "").strip(),

        # Date settings
        "DATE_MODE": os.getenv("DATE_MODE", "PREVIOUS_MONTH").strip().upper(),
        "DATE_FROM": os.getenv("DATE_FROM", "").strip(),
        "DATE_TO": os.getenv("DATE_TO", "").strip(),

        # VAT settings
        "VAT_ROUNDING_MODE": os.getenv("VAT_ROUNDING_MODE", "PER_TRANSACTION").strip().upper(),

        # Output settings
        "OUTPUT_DIR": os.getenv("OUTPUT_DIR", "output").strip(),
        "LOG_DIR": os.getenv("LOG_DIR", "logs").strip(),

        # SQL query file
        "SQL_QUERY_FILE": os.getenv("SQL_QUERY_FILE", "sql/query_daily_esales.sql").strip(),
    }

    required_fields = [
        "SQL_SERVER",
        "SQL_DATABASE",
        "SQL_USERNAME",
        "SQL_PASSWORD",
        "POS_BRANCH",
        "POS_COMPANY_NAME",
        "POS_OPERATOR_NAME",
        "POS_SERIAL_NO",
        "POS_MIN",
        "POS_PERMIT_NO",
    ]

    missing = [key for key in required_fields if not settings[key]]

    if missing:
        raise ValueError(
            "Missing required .env value(s): " + ", ".join(missing)
        )

    valid_date_modes = ["PREVIOUS_MONTH", "CURRENT_MONTH", "CUSTOM"]
    if settings["DATE_MODE"] not in valid_date_modes:
        raise ValueError(
            "Invalid DATE_MODE. Use PREVIOUS_MONTH, CURRENT_MONTH, or CUSTOM."
        )

    valid_vat_modes = ["PER_TRANSACTION", "SUMMARY"]
    if settings["VAT_ROUNDING_MODE"] not in valid_vat_modes:
        raise ValueError(
            "Invalid VAT_ROUNDING_MODE. Use PER_TRANSACTION or SUMMARY."
        )

    return settings


# ============================================================
# LOGGING
# ============================================================

def setup_logging(settings: dict) -> None:
    """
    Setup log file and console logging.

    Log file sample:
        logs/daily_esales_20260603.log
    """
    log_dir = resolve_project_path(settings["LOG_DIR"], "logs")
    log_dir.mkdir(parents=True, exist_ok=True)

    log_file = log_dir / f"daily_esales_{datetime.now().strftime('%Y%m%d')}.log"

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        handlers=[
            logging.FileHandler(log_file, encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ],
        force=True,
    )


# ============================================================
# DATE HANDLING
# ============================================================

def parse_date_yyyy_mm_dd(value: str, field_name: str) -> date:
    """
    Parse YYYY-MM-DD date from .env.
    """
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except Exception as exc:
        raise ValueError(f"Invalid {field_name}. Use YYYY-MM-DD format.") from exc


def get_month_range(year: int, month: int) -> tuple[date, date, str]:
    """
    Return:
        start_date
        end_date_exclusive
        month_label

    end_date_exclusive means first day after the report month.

    Example:
        April 2026:
            start_date = 2026-04-01
            end_date_exclusive = 2026-05-01
            month_label = 2026-04
    """
    start_date = date(year, month, 1)

    if month == 12:
        end_date_exclusive = date(year + 1, 1, 1)
    else:
        end_date_exclusive = date(year, month + 1, 1)

    month_label = start_date.strftime("%Y-%m")
    return start_date, end_date_exclusive, month_label


def get_previous_month_range() -> tuple[date, date, str]:
    """
    Get previous month based on today's date.
    """
    today = date.today()

    if today.month == 1:
        return get_month_range(today.year - 1, 12)

    return get_month_range(today.year, today.month - 1)


def get_current_month_range() -> tuple[date, date, str]:
    """
    Get current month based on today's date.
    """
    today = date.today()
    return get_month_range(today.year, today.month)


def get_custom_date_range(settings: dict) -> tuple[date, date, str]:
    """
    Get custom date range from .env.

    DATE_FROM and DATE_TO are inclusive dates.
    For SQL query, we convert DATE_TO into end_date_exclusive.
    """
    if not settings["DATE_FROM"] or not settings["DATE_TO"]:
        raise ValueError(
            "DATE_FROM and DATE_TO are required when DATE_MODE=CUSTOM."
        )

    start_date = parse_date_yyyy_mm_dd(settings["DATE_FROM"], "DATE_FROM")
    end_date_inclusive = parse_date_yyyy_mm_dd(settings["DATE_TO"], "DATE_TO")

    if end_date_inclusive < start_date:
        raise ValueError("DATE_TO cannot be earlier than DATE_FROM.")

    end_date_exclusive = date.fromordinal(end_date_inclusive.toordinal() + 1)

    month_label = (
        f"{start_date.strftime('%Y%m%d')}_to_"
        f"{end_date_inclusive.strftime('%Y%m%d')}"
    )

    return start_date, end_date_exclusive, month_label


def get_report_date_range(settings: dict) -> tuple[date, date, str]:
    """
    Decide report date range.

    Priority:
        1. Command line argument:
            py main.py 2026-04

        2. .env DATE_MODE:
            PREVIOUS_MONTH
            CURRENT_MONTH
            CUSTOM
    """
    if len(sys.argv) >= 2:
        month_text = sys.argv[1].strip()

        try:
            year, month = map(int, month_text.split("-"))
            return get_month_range(year, month)
        except Exception as exc:
            raise ValueError("Invalid month format. Use YYYY-MM. Example: 2026-04") from exc

    date_mode = settings["DATE_MODE"]

    if date_mode == "PREVIOUS_MONTH":
        return get_previous_month_range()

    if date_mode == "CURRENT_MONTH":
        return get_current_month_range()

    if date_mode == "CUSTOM":
        return get_custom_date_range(settings)

    raise ValueError("Invalid DATE_MODE.")


def get_end_date_inclusive(end_date_exclusive: date) -> date:
    """
    Convert exclusive end date into inclusive end date.

    Example:
        end_date_exclusive = 2026-05-01
        return = 2026-04-30
    """
    return date.fromordinal(end_date_exclusive.toordinal() - 1)


# ============================================================
# SQL CONNECTION
# ============================================================

def build_connection_string(settings: dict) -> str:
    """
    Build SQL Server connection string from .env.

    Notes:
        SQL_ENCRYPT and SQL_TRUST_SERVER_CERTIFICATE are included
        because ODBC Driver 18 may require encryption settings.
    """
    return (
        f"DRIVER={{{settings['SQL_DRIVER']}}};"
        f"SERVER={settings['SQL_SERVER']};"
        f"DATABASE={settings['SQL_DATABASE']};"
        f"UID={settings['SQL_USERNAME']};"
        f"PWD={settings['SQL_PASSWORD']};"
        f"Encrypt={settings['SQL_ENCRYPT']};"
        f"TrustServerCertificate={settings['SQL_TRUST_SERVER_CERTIFICATE']};"
    )


def connect_to_sql(settings: dict):
    """
    Open SQL Server connection using values from .env.
    """
    conn_str = build_connection_string(settings)

    logging.info("Connecting to SQL Server...")
    logging.info("SQL_SERVER: %s", settings["SQL_SERVER"])
    logging.info("SQL_DATABASE: %s", settings["SQL_DATABASE"])

    return pyodbc.connect(conn_str, timeout=30)


def read_sql_query(settings: dict) -> str:
    """
    Read SQL query from sql/query_daily_esales.sql.

    Why separate SQL file:
        Mas madaling baguhin ang query without touching Python code.
    """
    query_path = BASE_DIR / settings["SQL_QUERY_FILE"]

    if not query_path.exists():
        raise FileNotFoundError(f"SQL query file not found: {query_path}")

    return query_path.read_text(encoding="utf-8")


# ============================================================
# SQL PARAMETER CHECKING
# ============================================================

def count_sql_parameter_markers(query: str) -> int:
    """
    Count pyodbc parameter markers.

    pyodbc uses ? as parameter placeholder.

    Example:
        WHERE SaleDate >= ?
          AND SaleDate < ?

    This returns 2.

    Note:
        This simple count is enough for our current setup.
        Avoid putting question marks in SQL comments while testing.
    """
    return query.count("?")


# ============================================================
# DATA FETCHING
# ============================================================

def fetch_report_rows(settings: dict, start_date: date, end_date_exclusive: date) -> list[dict]:
    """
    Execute Daily E-Sales SQL query and return rows as dictionaries.

    Expected SQL output columns:
        LastInvoiceNo
        NetSalesWithVAT
        ZeroRated
        VATExempt
        Vatable
        TotalSalesNetOfVAT
        OutputVAT

    Static columns like TIN, POS Serial No., MIN, Permit No.
    are injected from .env after fetching.

    This function supports two SQL file modes:

    Mode 1:
        SQL has 0 parameter markers.

        Example:
            SELECT
                '000000' AS LastInvoiceNo,
                0.00 AS NetSalesWithVAT,
                0.00 AS ZeroRated,
                0.00 AS VATExempt,
                0.00 AS Vatable,
                0.00 AS TotalSalesNetOfVAT,
                0.00 AS OutputVAT;

        Python will run:
            cursor.execute(query)

    Mode 2:
        SQL has exactly 2 parameter markers.

        Example:
            WHERE ReceiptDate >= ?
              AND ReceiptDate < ?

        Python will run:
            cursor.execute(query, start_date, end_date_exclusive)
    """
    query = read_sql_query(settings)

    with connect_to_sql(settings) as conn:
        cursor = conn.cursor()

        logging.info("Running Daily E-Sales query...")
        logging.info("Report date from %s to before %s", start_date, end_date_exclusive)
        logging.info("VAT_ROUNDING_MODE: %s", settings["VAT_ROUNDING_MODE"])

        parameter_count = count_sql_parameter_markers(query)
        logging.info("SQL parameter markers found: %s", parameter_count)

        if parameter_count == 0:
            logging.warning(
                "SQL query has no ? parameter markers. Running query without date parameters."
            )
            cursor.execute(query)

        elif parameter_count == 2:
            cursor.execute(query, start_date, end_date_exclusive)

        else:
            raise ValueError(
                "SQL query must have either 0 or 2 parameter markers. "
                f"Found: {parameter_count}"
            )

        if cursor.description is None:
            raise ValueError(
                "SQL query did not return a result set. "
                "Make sure query_daily_esales.sql uses SELECT and returns report columns."
            )

        columns = [column[0] for column in cursor.description]
        rows = [dict(zip(columns, row)) for row in cursor.fetchall()]

    logging.info("Rows fetched: %s", len(rows))

    return rows


def apply_static_env_fields(settings: dict, rows: list[dict]) -> list[dict]:
    """
    Inject static POS/BIR details from .env into each report row.

    Why:
        These details usually come from POS permit/template and should
        be configurable per branch, not hardcoded in Python and not
        required from SQL.
    """
    enriched_rows = []

    for row in rows:
        new_row = dict(row)

        new_row["TIN"] = settings["POS_TIN"]
        new_row["TerminalNo"] = settings["POS_BRANCH"]
        new_row["POSSerialNo"] = settings["POS_SERIAL_NO"]
        new_row["MIN"] = settings["POS_MIN"]
        new_row["PermitNo"] = settings["POS_PERMIT_NO"]

        enriched_rows.append(new_row)

    return enriched_rows


# ============================================================
# VALUE CONVERSION
# ============================================================

def money(value) -> float:
    """
    Convert SQL numeric/decimal values to float for Excel.

    None becomes 0.00 to prevent blank total computations.
    """
    if value is None:
        return 0.0

    if isinstance(value, Decimal):
        return float(value)

    return float(value)


# ============================================================
# EXCEL STYLING
# ============================================================

def apply_basic_sheet_style(ws):
    """
    Apply layout and formatting based on APR.xls template.
    """
    widths = {
        "A": 14,
        "B": 14,
        "C": 24,
        "D": 26,
        "E": 30,
        "F": 16,
        "G": 18,
        "H": 14,
        "I": 14,
        "J": 16,
        "K": 22,
        "L": 14,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row_num in range(1, 5):
        ws.row_dimensions[row_num].height = 20

    ws.row_dimensions[7].height = 35

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9EAF7")

    for cell in ws[7]:
        if cell.column <= 12:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=8, max_row=ws.max_row, min_col=1, max_col=12):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col in range(7, 13):
        col_letter = get_column_letter(col)

        for row_num in range(8, ws.max_row + 1):
            ws[f"{col_letter}{row_num}"].number_format = "#,##0.00"


# ============================================================
# EXCEL REPORT GENERATION
# ============================================================

def build_excel_report(
    settings: dict,
    rows: list[dict],
    start_date: date,
    end_date_exclusive: date,
    month_label: str,
) -> Path:
    """
    Create Excel workbook using the APR.xls report layout.
    """
    if not rows:
        raise ValueError("No rows returned by SQL query. Report was not generated.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily E-Sales"

    end_date = get_end_date_inclusive(end_date_exclusive)

    # Header section based on APR.xls
    ws["A1"] = settings["POS_COMPANY_NAME"]
    ws["A2"] = settings["POS_OPERATOR_NAME"]
    ws["A3"] = "Daily E-Sales Report"
    ws["A4"] = f"For the Period of {start_date.strftime('%m/%d/%Y')} to {end_date.strftime('%m/%d/%Y')}"

    for cell_address in ["A1", "A2", "A3", "A4"]:
        ws[cell_address].font = Font(
            bold=True if cell_address in ["A1", "A3"] else False
        )

    headers = [
        "TIN",
        "Terminal No.",
        "POS Serial No.",
        "M.I.N",
        "Permit No.",
        "Last Invoice No.",
        "Net Sales (w/VAT)",
        "Zero Rated",
        "VAT Exempt",
        "Vatable",
        "Total Sales (Net of VAT)",
        "Output VAT",
    ]

    for col_index, header in enumerate(headers, start=1):
        ws.cell(row=7, column=col_index).value = header

    start_row = 8

    for row_index, row in enumerate(rows, start=start_row):
        ws.cell(row=row_index, column=1).value = row.get("TIN", "")
        ws.cell(row=row_index, column=2).value = row.get("TerminalNo", "")
        ws.cell(row=row_index, column=3).value = row.get("POSSerialNo", "")
        ws.cell(row=row_index, column=4).value = row.get("MIN", "")
        ws.cell(row=row_index, column=5).value = row.get("PermitNo", "")
        ws.cell(row=row_index, column=6).value = row.get("LastInvoiceNo", "")

        ws.cell(row=row_index, column=7).value = money(row.get("NetSalesWithVAT"))
        ws.cell(row=row_index, column=8).value = money(row.get("ZeroRated"))
        ws.cell(row=row_index, column=9).value = money(row.get("VATExempt"))
        ws.cell(row=row_index, column=10).value = money(row.get("Vatable"))
        ws.cell(row=row_index, column=11).value = money(row.get("TotalSalesNetOfVAT"))
        ws.cell(row=row_index, column=12).value = money(row.get("OutputVAT"))

    total_row = start_row + len(rows)

    ws.cell(row=total_row, column=1).value = "Grand Total :"
    ws.cell(row=total_row, column=1).font = Font(bold=True)

    for col in range(7, 13):
        col_letter = get_column_letter(col)
        ws.cell(row=total_row, column=col).value = (
            f"=SUM({col_letter}{start_row}:{col_letter}{total_row - 1})"
        )
        ws.cell(row=total_row, column=col).font = Font(bold=True)

    apply_basic_sheet_style(ws)
    ws.freeze_panes = "A8"

    output_dir = resolve_project_path(settings["OUTPUT_DIR"], "output")
    output_dir.mkdir(parents=True, exist_ok=True)

    branch = settings["POS_BRANCH"].replace(" ", "_").upper()
    output_file = output_dir / f"Daily_E-Sales_Report_{branch}_{month_label}.xlsx"

    wb.save(output_file)

    return output_file


# ============================================================
# MAIN PROCESS
# ============================================================

def main():
    try:
        settings = load_settings()
        setup_logging(settings)

        logging.info("Starting Daily E-Sales Report POS Automation...")

        start_date, end_date_exclusive, month_label = get_report_date_range(settings)

        logging.info("POS_BRANCH: %s", settings["POS_BRANCH"])
        logging.info("DATE_MODE: %s", settings["DATE_MODE"])
        logging.info("Report label: %s", month_label)

        rows = fetch_report_rows(settings, start_date, end_date_exclusive)
        rows = apply_static_env_fields(settings, rows)

        output_file = build_excel_report(
            settings=settings,
            rows=rows,
            start_date=start_date,
            end_date_exclusive=end_date_exclusive,
            month_label=month_label,
        )

        logging.info("Report generated successfully: %s", output_file)
        print(f"SUCCESS: Report generated: {output_file}")

    except Exception as exc:
        logging.exception("Failed to generate Daily E-Sales Report: %s", exc)
        print(f"ERROR: {exc}")
        sys.exit(1)


if __name__ == "__main__":
    main()